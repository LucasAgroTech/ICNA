from flask import Flask, render_template, redirect, url_for, flash, send_from_directory, Response
from reportlab.lib.pagesizes import landscape
from flask_sqlalchemy import SQLAlchemy
from flask_mail import Mail, Message
from flask_wtf import FlaskForm
from wtforms import StringField, SubmitField, SelectField, EmailField, BooleanField
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from wtforms.validators import DataRequired, Email, Regexp
from sqlalchemy.exc import IntegrityError
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from io import BytesIO
import qrcode
import os
import random
import string

app = Flask(__name__)
app.config['SECRET_KEY'] = 'secret_key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://cnivqiqnrnmrsu:542f156128aa2ebeca28cab03af01d727dadfdf04570e0954062774156cb6c03@ec2-34-235-108-214.compute-1.amazonaws.com:5432/dgfmsdkpr5tof'
db = SQLAlchemy(app)

app.config["MAIL_SERVER"] = "smtp.hostinger.com"
app.config["MAIL_PORT"] = 465
app.config["MAIL_USE_TLS"] = False
app.config["MAIL_USE_SSL"] = True
app.config["MAIL_USERNAME"] = os.environ["MAIL_USERNAME"]
app.config["MAIL_PASSWORD"] = os.environ["MAIL_PASSWORD"]

mail = Mail(app)

class Registration(db.Model):
    code = db.Column(db.String(8), unique=True, nullable=False)
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    cpf = db.Column(db.String(14), unique=True, nullable=False)
    email = db.Column(db.String(100), nullable=False)
    state = db.Column(db.String(50), nullable=False)
    city = db.Column(db.String(100), nullable=False)
    company_name = db.Column(db.String(100))
    role = db.Column(db.String(100))

class RegistrationForm(FlaskForm):
    name = StringField('Nome Completo', validators=[DataRequired()])
    cpf = StringField('CPF', validators=[
        DataRequired(),
        Regexp(r"^\d{3}\.\d{3}\.\d{3}-\d{2}$", message="Ajuste o CPF: xxx.xxx.xxx-xx")
    ])
    email = EmailField('E-mail', validators=[DataRequired(), Email()])
    state = SelectField('Estado', choices=[
        ('AC', 'Acre'),
        ('AL', 'Alagoas'),
        ('AP', 'Amapá'),
        ('AM', 'Amazonas'),
        ('BA', 'Bahia'),
        ('CE', 'Ceará'),
        ('DF', 'Distrito Federal'),
        ('ES', 'Espírito Santo'),
        ('GO', 'Goiás'),
        ('MA', 'Maranhão'),
        ('MT', 'Mato Grosso'),
        ('MS', 'Mato Grosso do Sul'),
        ('MG', 'Minas Gerais'),
        ('PA', 'Pará'),
        ('PB', 'Paraíba'),
        ('PR', 'Paraná'),
        ('PE', 'Pernambuco'),
        ('PI', 'Piauí'),
        ('RJ', 'Rio de Janeiro'),
        ('RN', 'Rio Grande do Norte'),
        ('RS', 'Rio Grande do Sul'),
        ('RO', 'Rondônia'),
        ('RR', 'Roraima'),
        ('SC', 'Santa Catarina'),
        ('SP', 'São Paulo'),
        ('SE', 'Sergipe'),
        ('TO', 'Tocantins')
    ])
    city = StringField('Cidade', validators=[DataRequired()]) 
    company_name = StringField('Empresa ou Instituição', validators=[DataRequired()])
    role = StringField('Cargo', validators=[DataRequired()])
    consent = BooleanField('Ao realizar minha inscrição, AUTORIZO que a Confederação da Agricultura e Pecuária do Brasil – CNA, sem contrapartidas ou ônus de qualquer natureza: (i) realize o tratamento de meus dados pessoais, em conformidade com o disposto na Lei nº 13.709, de 14 de agosto de 2018 (Lei Geral de Proteção de Dados - LGPD) e nos termos da Política de Privacidade do Sistema CNA/SENAR/ICNA (disponível em [http://www.cnabrasil.org.br)](https://www.cnabrasil.org.br/lei-geral-de-protecao-de-dados); (ii) reproduza, utilize e veicule meu nome, meus depoimentos/entrevistas – falados(as) ou escritos(as) – eventualmente concedidos(as), e minha imagem, relacionados ao evento e obtidos e/ou captados/gravados por qualquer meio e/ou forma durante sua realização, em materiais informativos e/ou publicitários, pesquisas, projetos e/ou campanhas institucionais próprios e/ou de qualquer entidade integrante do Sistema CNA/SENAR/ICNA, podendo divulgá-los ao público por qualquer meio, forma e veículo de comunicação, no território nacional e no exterior, e inclusive realizar cortes, reduções e (re)edições para utilização do material e, ainda, disponibilizar as imagens em aplicativos como o flickr, assim como podendo manter salvo o conteúdo para visualizações posteriores em mídias e/ou plataformas digitais, resguardando-me, porém, o direito de cancelar essas autorizações, parcial ou integralmente, a qualquer tempo e sem qualquer ônus, por meio do formulário disponível em [https://www.cnabrasil.org.br/lei-geral-de-protecao-de-dados](https://www.cnabrasil.org.br/lei-geral-de-protecao-de-dados).', validators=[DataRequired()])
    submit = SubmitField('Inscreva-se')

def generate_code():
    while True:
        code = "CNA-" + ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
        if not Registration.query.filter_by(code=code).first():
            return code

TICKET_SIZE = (3 * 72, 5 * 72)

def generate_pdf(name, cpf, company, code):
    file_path = f'static/pdfs/{code}.pdf'
    c = canvas.Canvas(file_path, pagesize=TICKET_SIZE)  # Mudança do pagesize para TICKET_SIZE
    width, height = TICKET_SIZE
    
    background_image_path = 'static/images/background.png'
    c.drawInlineImage(background_image_path, 0, 0, width=width, height=height)

    c.setFillColorRGB(1, 1, 1)
    font_size = 10
    spacing = 15
    
    y_position = height * 0.2
    
    # Centralizando o QR Code no meio
    qr_size = 110
    qr_x = (width - qr_size) / 2
    qr_y = (height - qr_size) / 2
    
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=1)
    qr.add_data(code)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img_path = f'static/qrcodes/{code}.png'
    img.save(img_path, 'PNG')
    c.drawInlineImage(img_path, qr_x, qr_y, width=qr_size, height=qr_size)
    
    # Adicionando o nome e o código abaixo do QR Code (ajuste conforme necessário)
    c.setFont("Helvetica", font_size)
    c.drawCentredString(width / 2, y_position, f"{name}")
    c.drawCentredString(width / 2, y_position - spacing, f"{cpf}")
    c.drawCentredString(width / 2, y_position - 2*spacing, f"{company}")
    c.save()
    return file_path

def send_email(recipient, subject, template, pdf_path=None, **kwargs):
    msg = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=[recipient])
    msg.html = render_template(template + '.html', **kwargs)
    if pdf_path:
        with app.open_resource(pdf_path) as fp:
            msg.attach(pdf_path.split("/")[-1], "application/pdf", fp.read())
    mail.send(msg)
    
def save_to_xlsx(data, filename=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Registrações"

    # Cabeçalhos
    columns = ["Código", "Nome", "CPF", "Email", "Estado", "Cidade", "Empresa", "Cargo"]
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = column_title
        ws.column_dimensions[col_letter].width = 15

    # Preenchendo os dados
    for row_num, registration in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=registration.code)
        ws.cell(row=row_num, column=2, value=registration.name)
        ws.cell(row=row_num, column=3, value=registration.cpf)
        ws.cell(row=row_num, column=4, value=registration.email)
        ws.cell(row=row_num, column=5, value=registration.state)
        ws.cell(row=row_num, column=6, value=registration.city)
        ws.cell(row=row_num, column=7, value=registration.company_name)
        ws.cell(row=row_num, column=8, value=registration.role)

    if filename:
        wb.save(filename)
        return filename
    else:
        temp_stream = BytesIO()
        wb.save(temp_stream)
        return temp_stream.getvalue()


@app.route('/view-registrations')
def view_registrations():
    registrations = Registration.query.all()
    return render_template('test.html', registrations=registrations)

@app.route('/', methods=['GET', 'POST'])
def index():
    form = RegistrationForm()
    pdf_link = None
    if form.validate_on_submit():
        name = form.name.data
        cpf = form.cpf.data
        company = form.company_name.data
        code = generate_code()
        
        registration = Registration(name=form.name.data,
            cpf=form.cpf.data,
            email=form.email.data,
            state=form.state.data,
            city=form.city.data,
            company_name=form.company_name.data,
            role=form.role.data,
            code=code)
        try:
            db.session.add(registration)
            db.session.commit()
            # Coloca a geração do PDF aqui, depois do commit.
            pdf_path = generate_pdf(name, cpf, company, code)
            
            send_email(
                recipient=form.email.data,
                subject="Confirmação de Inscrição",
                template="email_template",
                pdf_path=pdf_path,
                name=name,
                code=code
            )

            # Redirecionando diretamente para o download do PDF
            pdf_link = url_for('get_pdf', filename=code + ".pdf")
            
        except IntegrityError as e:
            db.session.rollback()
            if 'cpf' in str(e):
                flash('Este CPF já foi registrado.', 'error')
            else:
                flash('Ocorreu um erro ao registrar. Tente novamente.', 'error')
    
    return render_template('index.html', form=form, pdf_link=pdf_link)

@app.route('/pdfs/<filename>')
def get_pdf(filename):
    directory = os.path.join(app.root_path, 'static/pdfs')
    # Configurando a resposta para ser um arquivo a ser baixado
    response = send_from_directory(directory, filename, as_attachment=True)
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@app.route('/export-registrations-xlsx')
def export_registrations_xlsx():
    registrations = Registration.query.all()
    
    # Converta 'registrations' para uma lista de dicionários se ainda não for
    xlsx_data = save_to_xlsx(registrations, filename=None)  # Modifique a função save_to_xlsx para permitir None como filename
    
    response = Response(xlsx_data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = "attachment; filename=registrations.xlsx"
    
    return response


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)